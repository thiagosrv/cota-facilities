#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Converter XLSX para JSON
Uso: python converter_xlsx_to_json.py empresas.xlsx
"""

import openpyxl
import json
import sys
import io

# Force UTF-8 output
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def converter_xlsx_para_json(caminho_xlsx):
    """Converte planilha Excel em JSON estruturado."""

    wb = openpyxl.load_workbook(caminho_xlsx)
    dados = {}

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        # Remover caracteres especiais/encoding issues do nome da aba
        servico = nome_aba.strip()

        if servico not in dados:
            dados[servico] = {}

        # Ler cabeçalhos da primeira linha
        cabecalhos = []
        for celula in ws[1]:
            if celula.value:
                cabecalhos.append(celula.value.strip() if isinstance(celula.value, str) else celula.value)

        # Criar mapa de índices para os nomes de colunas esperados
        col_map = {}
        for idx, cab in enumerate(cabecalhos):
            cab_lower = str(cab).lower()
            # Mapear nomes portugueses para inglês
            if "cidade" in cab_lower:
                col_map["city"] = idx
            elif "nome" in cab_lower or "empresa" in cab_lower:
                col_map["name"] = idx
            elif "nota" in cab_lower or "rating" in cab_lower:
                col_map["rating"] = idx
            elif "endereço" in cab_lower or "endereco" in cab_lower or "address" in cab_lower:
                col_map["address"] = idx
            elif "telefone" in cab_lower or "phone" in cab_lower or "fone" in cab_lower or "tel" in cab_lower:
                col_map["phone"] = idx
            elif "site" in cab_lower or "website" in cab_lower or "web" in cab_lower:
                col_map["website"] = idx
            elif "link" in cab_lower or "maps" in cab_lower or "url" in cab_lower:
                col_map["url"] = idx

        # Agrupar por cidade
        cidades_dict = {}

        for linha_idx, linha in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            # Pular linhas vazias
            if all(celula.value is None for celula in linha):
                continue

            # Mapear valores usando o col_map
            registro = {}

            if "city" in col_map:
                cidade = linha[col_map["city"]].value
                if cidade:
                    registro["city"] = str(cidade).strip() if cidade else ""

            if "name" in col_map:
                name = linha[col_map["name"]].value
                registro["name"] = str(name).strip() if name else ""

            if "rating" in col_map:
                rating = linha[col_map["rating"]].value
                if rating:
                    try:
                        registro["rating"] = float(rating)
                    except:
                        registro["rating"] = 4.5
                else:
                    registro["rating"] = 4.5

            if "address" in col_map:
                address = linha[col_map["address"]].value
                registro["address"] = str(address).strip() if address else ""

            if "phone" in col_map:
                phone = linha[col_map["phone"]].value
                registro["phone"] = str(phone).strip() if phone else ""

            if "website" in col_map:
                website = linha[col_map["website"]].value
                val = str(website).strip() if website else ""
                # Garantir que começa com https://
                if val and not val.startswith("http"):
                    val = "https://" + val
                registro["website"] = val

            if "url" in col_map:
                url = linha[col_map["url"]].value
                # Se for "Ver no Maps" (texto sem link), substituir por vazio
                if url and "Ver no Maps" in str(url):
                    registro["url"] = ""
                else:
                    registro["url"] = str(url).strip() if url else ""

            # Agrupar por cidade
            cidade = registro.get("city", "")
            if cidade and registro.get("name"):
                if cidade not in cidades_dict:
                    cidades_dict[cidade] = []
                cidades_dict[cidade].append(registro)

        # Adicionar ao dicionário principal
        dados[servico].update(cidades_dict)

    # Salvar JSON na pasta js/ (onde o site busca)
    import os
    os.makedirs("js", exist_ok=True)
    with open("js/companies-db.json", "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

    print(f"✅ Convertido com sucesso: js/companies-db.json")
    print(f"📊 Serviços: {', '.join(dados.keys())}")
    for servico, cidades in dados.items():
        print(f"   {servico}: {len(cidades)} cidades")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python converter_xlsx_to_json.py [arquivo.xlsx]")
        sys.exit(1)

    caminho = sys.argv[1]
    converter_xlsx_para_json(caminho)
